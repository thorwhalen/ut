"""Reading excel (but more than just cell values: Also colors, formats, etc.)"""
import pandas as pd
import numpy as np
from openpyxl import load_workbook  # Get it here: https://pypi.org/project/openpyxl/ (or pip install openpyxl)


def get_sheet_data(xls_filepath, sheetname=None):
    wb = load_workbook(xls_filepath, data_only=True)
    sheetname = sheetname or next(iter(wb.sheetnames))
    sheet = wb[sheetname]
    return sheet


def pra(obj):
    print(*(x for x in dir(obj) if not x[0].startswith('_') and not x[0].isupper()), sep='\t')


def last_true_idx(series):
    return len(series) - series.iloc[::-1].values.argmax()


def mk_dotpath_getter(dotpath):
    attrs = dotpath.split('.')

    def attr_get(x):
        for a in attrs:
            x = getattr(x, a)
        return x

    return attr_get


def dfof(sheet, dotpath='value'):
    attr_get = mk_dotpath_getter(dotpath)
    data = [[attr_get(c) for c in row] for row in sheet]
    return pd.DataFrame(data)


def coord_val_and_bgcolor_of_cell(c):
    return {'coordinate': c.coordinate, 'value': c.value, 'rgb': c.fill.bgColor.rgb}


def crop(df, last_row_idx, last_col_idx):
    return df.iloc[:last_row_idx, :last_col_idx]


def matmask(df, mask):
    if isinstance(mask, pd.DataFrame):
        mask = mask.values
    d = df.values.copy()
    d[~mask] = ''
    return pd.DataFrame(d, index=df.index, columns=df.columns)


def row_lidx_to_mat_lidx(row_lidx, n_cols):
    return np.tile(row_lidx, (n_cols, 1)).T
