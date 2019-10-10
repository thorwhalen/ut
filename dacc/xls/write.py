__author__ = 'thor'

import os
import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.reader.excel import InvalidFileException
try:
    from xlwings import Workbook, Sheet
except ImportError as e:
    print(e)


def multiple_dfs_to_multiple_sheets(df_list, xls_filepath, sheet_list=None, **kwargs):
    """
    Writes multiple dataframes in different excel sheets.
    Input:
        * xls_filepath: The excel file to write into
        * And then there's several choices:
            * df_list (a list of dataframes) and sheet_list (a list of corresponding names)
            * df_list = a list of {sheet_name: dataframe}
            * df_list = a list of (sheet_name, dataframe) tuples, when the order of the sheets matters)
    --> If no sheet names are given, the function either gives the name of the dataframe (if any), or
    simply iterates over sheet numbers...
    """
    if sheet_list is None:
        if isinstance(df_list, dict):
            # df_list, sheet_list = zip(df_list.values(), df_list.keys())
            df_list, sheet_list = list(df_list.values()), list(df_list.keys())

        elif isinstance(df_list[0], tuple):
            sheet_list = [x[0] for x in df_list]
            df_list = [x[1] for x in df_list]
        else:
            sheet_list = []
            for i, df in enumerate(df_list):
                name = df.name
                if not name:
                    name = "sheet {}".format(i)
                sheet_list.append(name)

    writer = ExcelWriter(xls_filepath)
    for df, sheet_name in zip(df_list, sheet_list):
        df.to_excel(writer, sheet_name, **kwargs)

    writer.save()


def df_to_excel_without_overwriting_it(df, xls_filepath, sheet_name, **kwargs):
    """
    write df to an excel sheet without overwriting the whole excel file if it exists
    (may need to create the excel with some data in it already for this to work)
    """
    try:
        book = load_workbook(xls_filepath)
        writer = pd.ExcelWriter(xls_filepath, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        try:
            df.to_excel(excel_writer=writer, sheet_name=sheet_name, **kwargs)
        except TypeError:
            df = _replace_non_numeric_non_strings_with_strings(df)
            df.to_excel(excel_writer=writer, sheet_name=sheet_name, **kwargs)
        writer.save()
    except InvalidFileException:
        try:
            df.to_excel(excel_writer=xls_filepath, sheet_name=sheet_name, **kwargs)
        except TypeError:
            df = _replace_non_numeric_non_strings_with_strings(df)
            df.to_excel(excel_writer=writer, sheet_name=sheet_name, **kwargs)


def clear_sheet_contents_without_changing_formatting(xls_filepath, sheet_name):
    if os.path.exist(xls_filepath):  # else do nothing
        with Workbook(fullname=xls_filepath, app_visible=False) as wkb:
            Sheet(sheet=sheet_name, wkb=wkb).clear_contents()


def _replace_non_numeric_non_strings_with_strings(df):
    index_names = df.index.names
    df = df.reset_index(drop=False, inplace=False)
    for c in df.columns:
        if df[c].dtype.name == 'object':
            if not isinstance(df[c].iloc[0], str):
                df[c] = df[c].apply(str)
    df = df.set_index(index_names)
    return df

